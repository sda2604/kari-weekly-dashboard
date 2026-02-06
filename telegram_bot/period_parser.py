"""
Утилита для извлечения периода отчёта из Excel файлов
"""

import re
import sys
from pathlib import Path
from datetime import datetime
import openpyxl


def safe_print(text):
    """
    Безопасный вывод текста с обработкой UnicodeEncodeError
    
    Args:
        text (str): Текст для вывода
    
    Note:
        Используется для Windows консоли с проблемами кодировки
    """
    try:
        print(text)
    except UnicodeEncodeError:
        pass


def parse_period_from_filename(filename):
    """
    Извлекает период отчёта из названия файла используя regex
    
    Args:
        filename (str): Название файла (ex: "По регионам 12-18.01.xlsx")
    
    Returns:
        str | None: Период в формате "DD-DD месяц" или None
    
    Examples:
        >>> parse_period_from_filename("По регионам 12-18.01.xlsx")
        '12-18 января'
        >>> parse_period_from_filename("Рассылка 05-11 января.xlsx")
        '05-11 января'
    """
    # Паттерн: DD-DD.MM или DD-DD месяц
    patterns = [
        r'(\d{1,2})-(\d{1,2})\.(\d{1,2})',  # 12-18.01
        r'(\d{1,2})-(\d{1,2})\s+(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)',  # 12-18 января
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            if len(match.groups()) == 3 and '.' in pattern:
                # Формат: 12-18.01
                day_start, day_end, month = match.groups()
                month_names = {
                    '01': 'января', '02': 'февраля', '03': 'марта',
                    '04': 'апреля', '05': 'мая', '06': 'июня',
                    '07': 'июля', '08': 'августа', '09': 'сентября',
                    '10': 'октября', '11': 'ноября', '12': 'декабря'
                }
                month_name = month_names.get(month, month)
                return f"{day_start}-{day_end} {month_name}"
            else:
                # Формат: 12-18 января
                day_start, day_end, month_name = match.groups()
                return f"{day_start}-{day_end} {month_name}"
    
    return None


def parse_period_from_excel(filepath):
    """
    Извлекает период отчёта из содержимого Excel файла
    
    Просматривает первые 20 строк и 10 колонок в поисках паттернов дат.
    
    Args:
        filepath (Path): Путь к Excel файлу
    
    Returns:
        str | None: Период в формате "DD-DD месяц YYYY" или None
    
    Examples:
        Поддерживаемые форматы:
        - "12-18.01.2026" → "12-18 января 2026"
        - "12-18 января 2026" → "12-18 января 2026"
        - "12.01.2026 - 18.01.2026" → "12-18 января 2026"
    
    Note:
        Использует openpyxl в read_only режиме для производительности
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        
        # Проверяем первые 20 строк, первые 10 колонок
        for row in range(1, min(21, sheet.max_row + 1)):
            for col in range(1, min(11, sheet.max_column + 1)):
                cell_value = sheet.cell(row, col).value
                
                if cell_value and isinstance(cell_value, str):
                    # Ищем паттерны дат
                    patterns = [
                        r'(\d{1,2})-(\d{1,2})\.(\d{1,2})\.(\d{4})',  # 12-18.01.2026
                        r'(\d{1,2})-(\d{1,2})\s+(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)\s+(\d{4})',  # 12-18 января 2026
                        r'(\d{1,2})\.(\d{1,2})\.(\d{4})\s*-\s*(\d{1,2})\.(\d{1,2})\.(\d{4})',  # 12.01.2026 - 18.01.2026
                    ]
                    
                    for pattern in patterns:
                        match = re.search(pattern, cell_value, re.IGNORECASE)
                        if match:
                            groups = match.groups()
                            
                            if len(groups) == 4 and '.' in pattern and 'января' not in pattern:
                                # Формат: 12-18.01.2026
                                day_start, day_end, month, year = groups
                                month_names = {
                                    '01': 'января', '02': 'февраля', '03': 'марта',
                                    '04': 'апреля', '05': 'мая', '06': 'июня',
                                    '07': 'июля', '08': 'августа', '09': 'сентября',
                                    '10': 'октября', '11': 'ноября', '12': 'декабря'
                                }
                                month_name = month_names.get(month, month)
                                return f"{day_start}-{day_end} {month_name} {year}"
                            
                            elif len(groups) == 4 and 'января' in pattern.lower():
                                # Формат: 12-18 января 2026
                                day_start, day_end, month_name, year = groups
                                return f"{day_start}-{day_end} {month_name} {year}"
                            
                            elif len(groups) == 6:
                                # Формат: 12.01.2026 - 18.01.2026
                                day_start, month_start, year_start, day_end, month_end, year_end = groups
                                month_names = {
                                    '01': 'января', '02': 'февраля', '03': 'марта',
                                    '04': 'апреля', '05': 'мая', '06': 'июня',
                                    '07': 'июля', '08': 'августа', '09': 'сентября',
                                    '10': 'октября', '11': 'ноября', '12': 'декабря'
                                }
                                month_name = month_names.get(month_start, month_start)
                                return f"{day_start}-{day_end} {month_name} {year_start}"
        
        wb.close()
    except Exception as e:
        safe_print(f"Oshibka chtenija Excel: {e}")
    
    return None


def get_report_period():
    """
    Определяет период отчёта из еженедельных Excel файлов KARI
    
    Алгоритм поиска (по приоритету):
        1. Парсинг названия файла "По регионам.xlsx"
        2. Парсинг содержимого "По регионам.xlsx"
        3. Парсинг названия "Рассылка аксессуары.xlsx"
        4. Парсинг содержимого "Рассылка аксессуары.xlsx"
        5. Парсинг файла оборачиваемости
        6. Fallback: текущая дата
    
    Returns:
        str: Период отчёта в формате "DD-DD месяц YYYY" или "DD.MM.YYYY"
    
    Examples:
        >>> get_report_period()
        '12-18 января 2026'
    
    Note:
        Используется в generate_dashboard.py и send_dashboard.py
        для автоматического определения периода отчётов
    """
    script_dir = Path(__file__).parent
    input_dir = script_dir.parent / "input"
    
    # Список файлов для проверки (в порядке приоритета)
    files_to_check = [
        input_dir / "Отчет по приросту регионы" / "По регионам.xlsx",
        input_dir / "Отчет по приросту аксессуаров по магазинам" / "Рассылка аксессуары магазины.xlsx",
        input_dir / "Обувь остатки и оборачиваемость по группам товара" / "Отчет по оборачиваемости ТЗ регион ННВ.xlsx",
    ]
    
    # 1. Пробуем извлечь из названия файла
    for filepath in files_to_check:
        if filepath.exists():
            period = parse_period_from_filename(filepath.name)
            if period:
                safe_print(f"[OK] Period najden v nazvanii fajla: {period}")
                return period
    
    # 2. Пробуем извлечь из содержимого Excel
    for filepath in files_to_check:
        if filepath.exists():
            safe_print(f"[Poisk] Proverjaju soderzhimoe: {filepath.name}")
            period = parse_period_from_excel(filepath)
            if period:
                safe_print(f"[OK] Period najden v soderzhimom fajla: {period}")
                return period
    
    # 3. Используем текущую дату
    now = datetime.now()
    period = now.strftime("%d.%m.%Y")
    safe_print(f"[Info] Period ne najden, ispol'zuetsja tekuschaja data: {period}")
    return period


if __name__ == '__main__':
    # Тест
    period = get_report_period()
    safe_print(f"\n[Rezul'tat] Period otcheta: {period}")
