#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Система архивации входных Excel файлов
========================================
Автоматическая архивация файлов из input/ по неделям.

Логика:
1. Читает период из Excel файлов
2. Перемещает файлы: input/ → input/current/ (текущая неделя)
3. Архивирует старые: input/current/ → input/archive/YYYY-MM-DD_week/
4. Удаляет архивы старше 6 недель

Структура:
input/
├── current/                  # Текущая неделя (активные файлы)
│   ├── Отчет по приросту регионы/
│   ├── Отчет по приросту аксессуаров по магазинам/
│   └── Обувь остатки и оборачиваемость по группам товара/
└── archive/                  # Архив по неделям
    ├── 2026-01-12_week/      # Неделя 12-18 января
    ├── 2026-01-05_week/      # Неделя 5-11 января
    └── ...

Автор: Claude Code для KARI
Версия: 1.0
Дата: 26.01.2026
"""

import os
import sys
import shutil
from pathlib import Path
from datetime import datetime, timedelta
import re

# Настройка кодировки для Windows
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Пути
BASE_DIR = Path(__file__).parent.parent
INPUT_DIR = BASE_DIR / "input"
CURRENT_DIR = INPUT_DIR / "current"
ARCHIVE_DIR = INPUT_DIR / "archive"

# Константы
MAX_WEEKS_TO_KEEP = 6  # Хранить максимум 6 недель
FOLDERS_TO_ARCHIVE = [
    "Отчет по приросту регионы",
    "Отчет по приросту аксессуаров по магазинам",
    "Обувь остатки и оборачиваемость по группам товара"
]


def log(msg, level="INFO"):
    """Логирование с временем"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"[{timestamp}] [{level}] {msg}")


def parse_period_from_excel(filepath):
    """
    Извлекает период из Excel файла
    Возвращает дату начала недели в формате YYYY-MM-DD
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active

        # Проверяем первые 20 строк, первые 10 колонок
        for row in range(1, min(21, sheet.max_row + 1)):
            for col in range(1, min(11, sheet.max_column + 1)):
                cell_value = sheet.cell(row, col).value

                if cell_value and isinstance(cell_value, str):
                    # Паттерн: DD-DD.MM.YYYY или DD.MM.YYYY - DD.MM.YYYY
                    patterns = [
                        r'(\d{1,2})-\d{1,2}\.(\d{1,2})\.(\d{4})',  # 12-18.01.2026
                        r'(\d{1,2})\.(\d{1,2})\.(\d{4})\s*-',      # 12.01.2026 - ...
                    ]

                    for pattern in patterns:
                        match = re.search(pattern, cell_value)
                        if match:
                            groups = match.groups()
                            if len(groups) == 3:
                                day, month, year = groups
                                # Возвращаем дату начала недели
                                return f"{year}-{month.zfill(2)}-{day.zfill(2)}"

        wb.close()
    except Exception as e:
        log(f"Ошибка чтения Excel {filepath.name}: {e}", "WARN")

    return None


def get_week_folder_name(folder_path):
    """
    Определяет имя архивной папки по содержимому Excel файлов
    Возвращает: YYYY-MM-DD_week (дата начала недели)
    """
    # Ищем Excel файлы в папке
    excel_files = list(folder_path.glob("*.xlsx")) + list(folder_path.glob("*.xls"))

    for excel_file in excel_files:
        period = parse_period_from_excel(excel_file)
        if period:
            log(f"  ✓ Период определён из {excel_file.name}: {period}")
            return f"{period}_week"

    # Fallback: используем дату модификации - 7 дней
    log(f"  ⚠ Период не найден в Excel, используем дату модификации", "WARN")
    if excel_files:
        mod_time = datetime.fromtimestamp(excel_files[0].stat().st_mtime)
        week_start = mod_time - timedelta(days=7)
        return week_start.strftime("%Y-%m-%d") + "_week"

    # Последний fallback: текущая дата - 7 дней
    week_start = datetime.now() - timedelta(days=7)
    return week_start.strftime("%Y-%m-%d") + "_week"


def create_directory_structure():
    """Создаёт структуру папок если не существует"""
    log("Проверка структуры папок...")

    # Создаём current/
    CURRENT_DIR.mkdir(exist_ok=True)
    log(f"  ✓ {CURRENT_DIR.relative_to(BASE_DIR)}")

    # Создаём archive/
    ARCHIVE_DIR.mkdir(exist_ok=True)
    log(f"  ✓ {ARCHIVE_DIR.relative_to(BASE_DIR)}")


def move_to_current():
    """
    Перемещает новые файлы из input/ в input/current/
    Если в current/ уже есть файлы с тем же периодом - перезаписывает
    """
    log("\nШаг 1: Перемещение новых файлов в current/")

    moved_count = 0

    for folder_name in FOLDERS_TO_ARCHIVE:
        source_folder = INPUT_DIR / folder_name

        if not source_folder.exists():
            log(f"  ⊘ Пропускаю {folder_name} (не существует)")
            continue

        dest_folder = CURRENT_DIR / folder_name
        dest_folder.mkdir(parents=True, exist_ok=True)

        # Перемещаем все файлы из source → dest
        files_in_source = list(source_folder.glob("*"))

        if not files_in_source:
            log(f"  ⊘ {folder_name}: нет файлов")
            continue

        for file in files_in_source:
            if file.is_file():
                dest_file = dest_folder / file.name

                # Перезаписываем если существует
                if dest_file.exists():
                    dest_file.unlink()
                    log(f"  ↻ Перезаписан: {file.name}")
                else:
                    log(f"  → Перемещён: {file.name}")

                shutil.move(str(file), str(dest_file))
                moved_count += 1

    if moved_count > 0:
        log(f"✓ Перемещено файлов: {moved_count}")
    else:
        log("  Нет новых файлов для перемещения")


def archive_old_current():
    """
    Архивирует текущие файлы из current/ в archive/YYYY-MM-DD_week/
    Только если прошла неделя
    """
    log("\nШаг 2: Архивация старых файлов из current/")

    # Проверяем есть ли папки в current/
    folders_in_current = [f for f in CURRENT_DIR.iterdir() if f.is_dir()]

    if not folders_in_current:
        log("  Нет папок в current/ для архивации")
        return

    # Определяем имя архивной папки по первой найденной папке
    first_folder = folders_in_current[0]
    archive_name = get_week_folder_name(first_folder)
    archive_path = ARCHIVE_DIR / archive_name

    # Проверяем: если архив с таким именем уже существует - не дублируем
    if archive_path.exists():
        log(f"  ⊘ Архив {archive_name} уже существует (пропускаю)")
        return

    # Создаём архивную папку
    archive_path.mkdir(parents=True, exist_ok=True)
    log(f"  ✓ Создана архивная папка: {archive_name}")

    # Перемещаем все папки из current/ в archive/
    archived_count = 0

    for folder in folders_in_current:
        dest_folder = archive_path / folder.name

        log(f"  → Архивирую: {folder.name}")
        shutil.move(str(folder), str(dest_folder))
        archived_count += 1

    if archived_count > 0:
        log(f"✓ Заархивировано папок: {archived_count}")


def cleanup_old_archives():
    """Удаляет архивы старше MAX_WEEKS_TO_KEEP недель"""
    log(f"\nШаг 3: Очистка архивов старше {MAX_WEEKS_TO_KEEP} недель")

    if not ARCHIVE_DIR.exists():
        log("  Архивов нет")
        return

    # Получаем все архивные папки
    archive_folders = sorted([f for f in ARCHIVE_DIR.iterdir() if f.is_dir()])

    if len(archive_folders) <= MAX_WEEKS_TO_KEEP:
        log(f"  Архивов: {len(archive_folders)} (лимит {MAX_WEEKS_TO_KEEP}) - удаление не требуется")
        return

    # Удаляем самые старые
    folders_to_delete = archive_folders[:-MAX_WEEKS_TO_KEEP]

    for folder in folders_to_delete:
        log(f"  🗑 Удаляю старый архив: {folder.name}")
        shutil.rmtree(folder)

    log(f"✓ Удалено архивов: {len(folders_to_delete)}")


def show_structure():
    """Показывает текущую структуру папок"""
    log("\n" + "="*60)
    log("ТЕКУЩАЯ СТРУКТУРА:")
    log("="*60)

    # Current
    log("\n📂 input/current/ (активные файлы):")
    if CURRENT_DIR.exists():
        current_folders = [f for f in CURRENT_DIR.iterdir() if f.is_dir()]
        if current_folders:
            for folder in current_folders:
                file_count = len(list(folder.glob("*")))
                log(f"  ├── {folder.name}/ ({file_count} файлов)")
        else:
            log("  └── (пусто)")
    else:
        log("  └── (не существует)")

    # Archive
    log("\n📦 input/archive/ (история по неделям):")
    if ARCHIVE_DIR.exists():
        archive_folders = sorted([f for f in ARCHIVE_DIR.iterdir() if f.is_dir()], reverse=True)
        if archive_folders:
            for i, folder in enumerate(archive_folders):
                prefix = "├──" if i < len(archive_folders) - 1 else "└──"
                log(f"  {prefix} {folder.name}/")
        else:
            log("  └── (пусто)")
    else:
        log("  └── (не существует)")

    log("\n" + "="*60)


def main():
    """Главная функция"""
    log("="*60)
    log("АРХИВАЦИЯ ВХОДНЫХ ФАЙЛОВ KARI")
    log("="*60)

    try:
        # 1. Создаём структуру
        create_directory_structure()

        # 2. Перемещаем новые файлы в current/
        move_to_current()

        # 3. Архивируем старые файлы из current/
        archive_old_current()

        # 4. Удаляем старые архивы
        cleanup_old_archives()

        # 5. Показываем результат
        show_structure()

        log("\n✅ АРХИВАЦИЯ ЗАВЕРШЕНА УСПЕШНО!")
        return 0

    except Exception as e:
        log(f"\n❌ ОШИБКА: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
